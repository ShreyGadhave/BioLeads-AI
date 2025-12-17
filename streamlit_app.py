"""
ðŸ§¬ BioLeads AI - 3D In-Vitro Lead Scoring Dashboard

An AI-powered lead qualification tool for researchers in drug toxicology.
Now with LIVE data from PubMed and NIH!
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
import os
import sys
import io
from pathlib import Path
from datetime import datetime

# Add app to path for imports
sys.path.insert(0, str(Path(__file__).parent))

# Page config - must be first Streamlit command
st.set_page_config(
    page_title="BioLeads AI - Lead Scoring",
    page_icon="ðŸ§¬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for premium dark mode UI
st.markdown("""
<style>
    /* Main app styling */
    .stApp {
        background: linear-gradient(135deg, #0f0f23 0%, #1a1a2e 50%, #16213e 100%);
    }
    
    /* Header styling */
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
    }
    
    .sub-header {
        color: #94a3b8;
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    
    /* Metric cards with glassmorphism */
    .metric-card {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 16px;
        padding: 1.5rem;
        text-align: center;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 40px rgba(102, 126, 234, 0.2);
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 700;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .metric-label {
        color: #94a3b8;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-top: 0.5rem;
    }
    
    /* Score badges */
    .score-hot {
        background: linear-gradient(90deg, #f43f5e 0%, #ec4899 100%);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.85rem;
    }
    
    .score-high {
        background: linear-gradient(90deg, #f59e0b 0%, #fbbf24 100%);
        color: #1a1a2e;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.85rem;
    }
    
    .score-medium {
        background: linear-gradient(90deg, #3b82f6 0%, #60a5fa 100%);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.85rem;
    }
    
    .score-low {
        background: rgba(148, 163, 184, 0.2);
        color: #94a3b8;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.85rem;
    }
    
    /* Sidebar styling */
    .css-1d391kg {
        background: rgba(15, 15, 35, 0.95);
    }
    
    /* Table hover effect */
    .dataframe tbody tr:hover {
        background: rgba(102, 126, 234, 0.1) !important;
    }
    
    /* Button styling */
    .stButton > button {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: scale(1.05);
        box-shadow: 0 10px 30px rgba(102, 126, 234, 0.3);
    }
    
    /* Slider styling */
    .stSlider > div > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Info boxes */
    .info-box {
        background: rgba(102, 126, 234, 0.1);
        border-left: 4px solid #667eea;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    
    /* Live indicator */
    .live-indicator {
        display: inline-block;
        width: 8px;
        height: 8px;
        background: #22c55e;
        border-radius: 50%;
        margin-right: 8px;
        animation: pulse 2s infinite;
    }
    
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.5; }
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# === DATA SOURCES ===

@st.cache_data(ttl=3600)  # Cache for 1 hour
def fetch_pubmed_leads(max_results=30):
    """Fetch real leads from PubMed."""
    import requests
    import xml.etree.ElementTree as ET
    
    ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    
    # Search for DILI/hepatotoxicity papers
    search_query = '("drug-induced liver injury" OR "DILI" OR "hepatotoxicity" OR "3D liver model") AND ("2023"[pdat] OR "2024"[pdat])'
    
    try:
        # Search PubMed
        search_resp = requests.get(ESEARCH_URL, params={
            "db": "pubmed",
            "term": search_query,
            "retmax": max_results,
            "retmode": "json",
            "sort": "relevance"
        }, timeout=10)
        
        pmids = search_resp.json().get("esearchresult", {}).get("idlist", [])
        
        if not pmids:
            return []
        
        # Fetch details
        fetch_resp = requests.get(EFETCH_URL, params={
            "db": "pubmed",
            "id": ",".join(pmids[:20]),
            "retmode": "xml"
        }, timeout=15)
        
        root = ET.fromstring(fetch_resp.text)
        leads = []
        seen_names = set()
        
        for article in root.findall(".//PubmedArticle"):
            # Get title
            title_elem = article.find(".//ArticleTitle")
            title = title_elem.text if title_elem is not None else ""
            
            # Get authors
            for author in article.findall(".//Author"):
                last_name = author.find("LastName")
                first_name = author.find("ForeName")
                affiliation = author.find(".//Affiliation")
                
                if last_name is not None:
                    name = f"Dr. {first_name.text if first_name else ''} {last_name.text}".strip()
                    
                    if name in seen_names:
                        continue
                    seen_names.add(name)
                    
                    aff_text = affiliation.text if affiliation is not None else ""
                    company, location = parse_affiliation(aff_text)
                    
                    leads.append({
                        "name": name,
                        "title": "Research Author",
                        "company": company,
                        "company_type": "Academic" if any(x in company.lower() for x in ["university", "institute", "college"]) else "Industry",
                        "location": location,
                        "hq_location": location,
                        "email": "",
                        "linkedin": "",
                        "funding_status": "Unknown",
                        "recent_publication": f"{title[:80]}..." if len(title) > 80 else title,
                        "publication_keywords": ["DILI", "hepatotoxicity"],
                        "uses_invitro": True,
                        "source": "PubMed"
                    })
                    break  # One author per paper
        
        return leads
    except Exception as e:
        st.warning(f"PubMed fetch error: {e}")
        return []


@st.cache_data(ttl=3600)
def fetch_nih_grants_leads(max_results=20):
    """Fetch real leads from NIH RePORTER."""
    import requests
    
    NIH_URL = "https://api.reporter.nih.gov/v2/projects/search"
    
    try:
        payload = {
            "criteria": {
                "advanced_text_search": {
                    "operator": "and",
                    "search_field": "all",
                    "search_text": "drug induced liver injury OR hepatotoxicity OR 3D liver model"
                },
                "fiscal_years": [2023, 2024, 2025],
                "include_active_projects": True
            },
            "offset": 0,
            "limit": max_results
        }
        
        resp = requests.post(NIH_URL, json=payload, timeout=15)
        results = resp.json().get("results", [])
        
        leads = []
        seen_names = set()
        
        for grant in results:
            pis = grant.get("principal_investigators", [])
            if not pis:
                continue
            
            pi = pis[0]
            name = f"Dr. {pi.get('first_name', '')} {pi.get('last_name', '')}".strip()
            
            if name in seen_names or name == "Dr. ":
                continue
            seen_names.add(name)
            
            org = grant.get("organization", {})
            award = grant.get("award_amount", 0)
            
            leads.append({
                "name": name,
                "title": "Principal Investigator",
                "company": org.get("org_name", "Unknown"),
                "company_type": "Academic",
                "location": f"{org.get('org_city', '')}, {org.get('org_state', '')}",
                "hq_location": f"{org.get('org_city', '')}, {org.get('org_state', '')}",
                "email": "",
                "linkedin": "",
                "funding_status": f"NIH Grant (${award:,})" if award else "NIH Grant",
                "recent_publication": grant.get("project_title", "")[:80],
                "publication_keywords": ["NIH", "grant"],
                "uses_invitro": True,
                "source": "NIH RePORTER"
            })
        
        return leads
    except Exception as e:
        st.warning(f"NIH fetch error: {e}")
        return []


def parse_affiliation(aff: str):
    """Parse affiliation string."""
    if not aff:
        return "Unknown Institution", "Unknown"
    
    parts = aff.split(",")
    company = parts[0].strip()[:50] if parts else "Unknown"
    
    location = "Unknown"
    for part in reversed(parts):
        part = part.strip()
        if any(x in part.lower() for x in ["usa", "uk", "germany", "ma", "ca", "ny", "tx"]):
            location = part[:30]
            break
    
    return company, location


def load_sample_leads():
    """Load sample leads from JSON file."""
    data_path = Path(__file__).parent / "data" / "sample_leads.json"
    
    if data_path.exists():
        with open(data_path, "r") as f:
            return json.load(f)
    return []


def calculate_score(lead):
    """Calculate probability score for a lead."""
    score = 0
    
    # Scientific Intent (+40)
    publication = lead.get("recent_publication", "") or ""
    scientific_keywords = ["dili", "liver", "hepatic", "hepato", "3d", "organoid", "spheroid", "toxicity"]
    if any(kw in publication.lower() for kw in scientific_keywords):
        score += 40
    
    # Role Fit (+30)
    title = lead.get("title", "").lower()
    role_keywords = ["toxicology", "toxicologist", "safety", "preclinical", "hepatic", "director", "vp", "head", "principal"]
    if any(kw in title for kw in role_keywords):
        score += 30
    
    # Company Intent (+20)
    funding = lead.get("funding_status", "").lower()
    if "series a" in funding or "series b" in funding:
        score += 20
    elif "series c" in funding or "public" in funding or "nih" in funding:
        score += 15
    
    # Technographic (+15)
    if lead.get("uses_invitro", False):
        score += 15
    
    # Location (+10)
    location = (lead.get("location", "") + " " + lead.get("hq_location", "")).lower()
    hub_locations = ["boston", "cambridge", "san francisco", "bay area", "basel", "london", "uk", "palo alto"]
    if any(hub in location for hub in hub_locations):
        score += 10
    
    return min(score, 100)


def get_score_tier(score):
    """Get tier label for a score."""
    if score >= 80:
        return "ðŸ”¥ Hot Lead"
    elif score >= 60:
        return "â­ High Priority"
    elif score >= 40:
        return "ðŸ“Š Medium"
    elif score >= 20:
        return "ðŸ“‰ Low"
    else:
        return "â„ï¸ Cold"


def generate_excel_export(df):
    """Generate formatted Excel file for export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Prepare export dataframe with proper columns
    export_cols = ["rank", "probability_score", "name", "title", "company", "location", "hq_location", "funding_status", "recent_publication", "source"]
    available_cols = [col for col in export_cols if col in df.columns]
    export_df = df[available_cols].copy()
    
    # Add Tier column
    if "probability_score" in export_df.columns:
        export_df.insert(1, "tier", export_df["probability_score"].apply(get_score_tier))
    
    # Rename columns to proper display names
    column_names = {
        "rank": "Rank",
        "tier": "Tier",
        "probability_score": "Probability Score",
        "name": "Name",
        "title": "Title",
        "company": "Company",
        "location": "Location",
        "hq_location": "HQ Location",
        "funding_status": "Funding Status",
        "recent_publication": "Recent Publication",
        "source": "Data Source"
    }
    export_df = export_df.rename(columns=column_names)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BioLeads Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Tier colors
    tier_fills = {
        "ðŸ”¥ Hot Lead": PatternFill(start_color="F43F5E", end_color="F43F5E", fill_type="solid"),
        "â­ High Priority": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
        "ðŸ“Š Medium": PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid"),
        "ðŸ“‰ Low": PatternFill(start_color="94A3B8", end_color="94A3B8", fill_type="solid"),
        "â„ï¸ Cold": PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
    }
    
    # Write headers
    for col_idx, column in enumerate(export_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Apply tier coloring
            col_name = export_df.columns[col_idx - 1]
            if col_name == "Tier" and value in tier_fills:
                cell.fill = tier_fills[value]
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = center_alignment
            elif col_name in ["Rank", "Probability Score"]:
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
    
    # Set column widths
    column_widths = {
        "Rank": 8,
        "Tier": 18,
        "Probability Score": 18,
        "Name": 25,
        "Title": 30,
        "Company": 35,
        "Location": 25,
        "HQ Location": 25,
        "Funding Status": 25,
        "Recent Publication": 50,
        "Data Source": 15
    }
    
    for col_idx, column in enumerate(export_df.columns, 1):
        width = column_widths.get(column, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_data = [
        ["BioLeads AI - Export Summary", ""],
        ["", ""],
        ["Total Leads", len(export_df)],
        ["Hot Leads (80+)", len(export_df[export_df.get("Probability Score", 0) >= 80]) if "Probability Score" in export_df.columns else 0],
        ["High Priority (60-79)", len(export_df[(export_df.get("Probability Score", 0) >= 60) & (export_df.get("Probability Score", 0) < 80)]) if "Probability Score" in export_df.columns else 0],
        ["Average Score", round(export_df["Probability Score"].mean(), 1) if "Probability Score" in export_df.columns else 0],
        ["", ""],
        ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Generated by", "BioLeads AI - Lead Scoring Agent"]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="667EEA")
            elif col_idx == 1:
                cell.font = Font(bold=True)
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def main():
    # Header
    st.markdown('<h1 class="main-header">ðŸ§¬ 3D In-Vitro Lead Qualification Dashboard</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header"><span class="live-indicator"></span>Live data from PubMed & NIH RePORTER | AI-powered lead scoring for biotech</p>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("## ðŸŽ¯ Filters")
        
        # Data source selector
        data_source = st.radio(
            "ðŸ“¡ Data Source",
            ["Sample Data (25 leads)", "ðŸ”´ Live: PubMed + NIH"],
            index=0,
            help="Sample data loads instantly. Live data fetches from APIs."
        )
        
        # Search box
        search_term = st.text_input(
            "ðŸ” Search",
            placeholder="Name, company, location...",
            help="Filter leads by any text"
        )
        
        # Score slider
        min_score = st.slider(
            "ðŸ“Š Minimum Probability Score",
            min_value=0,
            max_value=100,
            value=0,
            step=5
        )
        
        st.markdown("---")
        st.markdown("### ðŸ“ˆ Scoring Model")
        st.markdown("""
        | Signal | Weight |
        |--------|--------|
        | Scientific Intent | +40 |
        | Role Fit | +30 |
        | Company Intent | +20 |
        | Technographic | +15 |
        | Location | +10 |
        """)
    
    # Load data based on selection
    if "Live" in data_source:
        with st.spinner("ðŸ”„ Fetching live data from PubMed & NIH..."):
            pubmed_leads = fetch_pubmed_leads(30)
            nih_leads = fetch_nih_grants_leads(20)
            leads_raw = pubmed_leads + nih_leads
            
            if not leads_raw:
                st.warning("No live data fetched. Using sample data.")
                leads_raw = load_sample_leads()
            else:
                st.success(f"âœ… Fetched {len(pubmed_leads)} from PubMed, {len(nih_leads)} from NIH")
    else:
        leads_raw = load_sample_leads()
    
    # Calculate scores
    for lead in leads_raw:
        if "probability_score" not in lead:
            if "scores" in lead:
                lead["probability_score"] = sum(lead["scores"].values())
            else:
                lead["probability_score"] = calculate_score(lead)
    
    # Sort and rank
    leads_raw.sort(key=lambda x: x["probability_score"], reverse=True)
    for i, lead in enumerate(leads_raw, 1):
        lead["rank"] = i
    
    # Create DataFrame
    df = pd.DataFrame(leads_raw)
    
    # Apply filters
    filtered_df = df.copy()
    
    if search_term:
        mask = filtered_df.apply(lambda row: search_term.lower() in str(row).lower(), axis=1)
        filtered_df = filtered_df[mask]
    
    filtered_df = filtered_df[filtered_df["probability_score"] >= min_score]
    
    # Metrics row
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(filtered_df)}</div>
            <div class="metric-label">Total Leads</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        hot_leads = len(filtered_df[filtered_df["probability_score"] >= 80])
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{hot_leads}</div>
            <div class="metric-label">Hot Leads ðŸ”¥</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        avg_score = filtered_df["probability_score"].mean() if len(filtered_df) > 0 else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{avg_score:.0f}</div>
            <div class="metric-label">Avg Score</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        sources = filtered_df["source"].nunique() if "source" in filtered_df.columns else 1
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{sources}</div>
            <div class="metric-label">Data Sources</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Tabs
    tab1, tab2, tab3 = st.tabs(["ðŸ“‹ Lead Table", "ðŸ“Š Score Distribution", "ðŸ—ºï¸ Analysis"])
    
    with tab1:
        col1, col2, col3 = st.columns([1, 1, 4])
        with col1:
            excel_data = generate_excel_export(filtered_df)
            st.download_button(
                "ðŸ“¥ Download Excel",
                excel_data,
                f"bioleads_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            if st.button("ðŸ”„ Refresh"):
                st.cache_data.clear()
                st.rerun()
        
        # Prepare display
        display_cols = ["rank", "probability_score", "name", "title", "company", "location", "hq_location", "source"]
        available_cols = [col for col in display_cols if col in filtered_df.columns]
        display_df = filtered_df[available_cols].copy()
        
        display_df = display_df.rename(columns={
            "rank": "Rank", "probability_score": "Probability",
            "name": "Name", "title": "Title", "company": "Company",
            "location": "Location", "hq_location": "HQ", "source": "Source"
        })
        
        display_df["Tier"] = display_df["Probability"].apply(get_score_tier)
        
        cols_order = ["Rank", "Tier", "Probability", "Name", "Title", "Company", "Location", "HQ", "Source"]
        cols_order = [c for c in cols_order if c in display_df.columns]
        display_df = display_df[cols_order]
        
        st.dataframe(
            display_df,
            use_container_width=True,
            height=500,
            column_config={
                "Probability": st.column_config.ProgressColumn("Probability", min_value=0, max_value=100, format="%d%%")
            }
        )
    
    with tab2:
        col1, col2 = st.columns(2)
        
        with col1:
            fig_hist = px.histogram(filtered_df, x="probability_score", nbins=20, color_discrete_sequence=["#667eea"], title="Score Distribution")
            fig_hist.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0")
            st.plotly_chart(fig_hist, use_container_width=True)
        
        with col2:
            tier_counts = filtered_df["probability_score"].apply(
                lambda x: "Hot (80+)" if x >= 80 else "High (60-79)" if x >= 60 else "Medium (40-59)" if x >= 40 else "Low (<40)"
            ).value_counts()
            
            fig_pie = px.pie(values=tier_counts.values, names=tier_counts.index, title="Lead Tier Breakdown",
                            color_discrete_sequence=["#f43f5e", "#f59e0b", "#3b82f6", "#6b7280"])
            fig_pie.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0")
            st.plotly_chart(fig_pie, use_container_width=True)
    
    with tab3:
        col1, col2 = st.columns(2)
        
        with col1:
            if "company_type" in filtered_df.columns:
                type_scores = filtered_df.groupby("company_type")["probability_score"].mean().sort_values(ascending=True)
                fig_bar = px.bar(x=type_scores.values, y=type_scores.index, orientation='h',
                               color=type_scores.values, color_continuous_scale=["#6b7280", "#667eea", "#f43f5e"],
                               title="Average Score by Company Type")
                fig_bar.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0", showlegend=False)
                st.plotly_chart(fig_bar, use_container_width=True)
        
        with col2:
            if "source" in filtered_df.columns:
                source_counts = filtered_df["source"].value_counts()
                fig_source = px.bar(x=source_counts.values, y=source_counts.index, orientation='h',
                                   color_discrete_sequence=["#764ba2"], title="Leads by Data Source")
                fig_source.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0")
                st.plotly_chart(fig_source, use_container_width=True)
    
    # Footer
    st.markdown("---")
    st.markdown(
        f"""<div style="text-align: center; color: #64748b; padding: 1rem;">
            <p><strong>BioLeads AI</strong> | 3D In-Vitro Lead Scoring Agent</p>
            <p style="font-size: 0.8rem;">Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        </div>""",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
